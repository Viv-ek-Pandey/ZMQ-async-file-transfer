import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def createWorkBook(path):
    """Creates workbook on the provided path"""
    wb = openpyxl.Workbook()
    wb.save(filename=path)


def getRowCount(filepath, sheetname):
    """get row count from Excel file"""
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheetname]
    return sheet.max_row


def getColCount(filename, sheetname):
    """get column count from Excel file"""
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return sheet.max_column


def readata(filename, sheetname, rowNum, colNum):
    """read excel cell data (filename=excel filename, sheet=sheetname,rowNum,column)"""
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return sheet.cell(row=rowNum, column=colNum).value


def writeData(filename, sheetname, rowNum, colNum, data):
    """write data in Excel cell"""
    workbook = openpyxl.load_workbook(filename)
    if sheetname not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheetname)
    else:
        sheet = workbook[sheetname]
    sheet.cell(row=rowNum, column=colNum).value = str(data)
    workbook.close()
    workbook.save(filename)


def getOrCreateSheet(filename, sheetname=None):
    """create a sheet in Excel file"""
    workbook = openpyxl.load_workbook(filename)
    if sheetname not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheetname)
        workbook.close()
        workbook.save(filename)
    else:
        sheet = workbook[sheetname]
    return sheet


def suiteReportHeader(filepath, ws, suiteHeader, testCaseName):
    rowCount = getRowCount(filepath, ws)
    writeData(filepath, ws, rowCount + 1, 1, '')
    writeData(filepath, ws, rowCount + 2, 1, testCaseName)
    for i, ind in enumerate(suiteHeader):
        writeData(filepath, ws, rowCount + 3, i + 1, ind)
    adjustColWidthBasedOntextLen(filepath, ws)


def adjustColWidthBasedOntextLen(filename, sheetname):
    """adjust column width in Excel report based on column value  length
        this should be used  after all the data has been fed to excel"""

    workbook = openpyxl.load_workbook(filename)
    if sheetname not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheetname)
    else:
        sheet = workbook[sheetname]
    column_widths = []
    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            except IndexError:
                column_widths.append(len(str(cell.value)))
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    for i, column_width in enumerate(column_widths):
        minVal = column_width * 3 if column_width < 10 else column_width + 10
        sheet.column_dimensions[get_column_letter(i + 1)].width = minVal

    workbook.save(filename)


def mergeCol(filename, sheetname, rs, re, cs, ce):
    """merge column and rows in Excel file"""
    workbook = openpyxl.load_workbook(filename)
    if sheetname not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheetname)
    else:
        sheet = workbook[sheetname]
    sheet.merge_cells(start_row=rs, start_column=cs, end_row=re, end_column=ce)


def addTestResultToReportSheet(filepath, worksheet, header, data, status):
    """For adding test case header in test suite work sheet"""

    rowCound = getRowCount(filepath, worksheet)
    for ind, title in enumerate(header):
        if ind == len(header) - 1:
            writeData(filepath, worksheet, rowCound + 1, ind + 1, status)
        else:
            title = title.split(' ')
            title = ''.join(title)
            titleLowered = title[0].lower() + title[1:]
            if isinstance(data, dict):
                titleVal = data[titleLowered]
            else:
                titleVal = data
            writeData(filepath, worksheet, rowCound + 1, ind + 1, titleVal)
    adjustColWidthBasedOntextLen(filepath, worksheet)
